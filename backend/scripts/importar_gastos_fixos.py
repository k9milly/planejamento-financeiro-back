"""Importa a tabela de gastos fixos da planilha original.

Uso:
    python -m scripts.importar_gastos_fixos "Planejamento.xlsx" --ano 2026 --simular

Na planilha, os gastos fixos ficam em uma tabela nomeada `GastosFixos` com as
colunas DESCRIÇÃO, VALOR, DATA, TIPO (forma de pagamento), SITUAÇÃO e SOMAR.
Ela existe em uma aba só, porque descreve despesas recorrentes do ano inteiro.

**Não gera lançamentos.** Na planilha esses gastos aparecem como "Pago", mas os
lançamentos correspondentes já vêm da importação principal — marcá-los como
pagos aqui criaria um segundo lançamento para o mesmo dinheiro. Os gastos
entram como modelos pendentes; marcar o que já foi pago é decisão de quem usa.
"""

from __future__ import annotations

import argparse
import sys
import unicodedata
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl

from app.database import SessionLocal, criar_tabelas
from app.models import Ano, Categoria, GastoFixo

OBRIGATORIAS = {"DESCRICAO", "VALOR", "DATA"}

# Todos os gastos fixos da planilha estão sob esta categoria nos lançamentos.
CATEGORIA_PADRAO = "Fixo"


def normalizar(texto: object) -> str:
    if texto is None:
        return ""
    sem_acento = unicodedata.normalize("NFKD", str(texto))
    sem_acento = "".join(c for c in sem_acento if not unicodedata.combining(c))
    return sem_acento.strip().upper()


def para_decimal(valor: object) -> Decimal | None:
    if valor is None:
        return None
    try:
        return Decimal(str(valor)).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return None


def localizar_tabela(ws) -> tuple[int, int, dict[str, int]] | None:
    """Acha a tabela de gastos fixos pelas colunas, não pela posição."""
    for nome in list(ws.tables):
        tabela = ws.tables[nome]
        colunas = [normalizar(c.name) for c in tabela.tableColumns]
        if not OBRIGATORIAS <= set(colunas):
            continue
        col_ini, linha_ini, _, linha_fim = openpyxl.utils.range_boundaries(tabela.ref)
        return (
            linha_ini + 1,
            linha_fim,
            {nome_col: col_ini + i for i, nome_col in enumerate(colunas)},
        )
    return None


def importar(caminho: Path, ano_numero: int, simular: bool) -> None:
    wb = openpyxl.load_workbook(caminho, data_only=True)
    sessao = SessionLocal()

    try:
        ano = sessao.query(Ano).filter(Ano.ano == ano_numero).one_or_none()
        if ano is None:
            print(
                f"ERRO: o ano {ano_numero} não existe. Rode a importação da "
                "planilha primeiro.",
                file=sys.stderr,
            )
            sys.exit(1)

        existentes = {
            normalizar(g.descricao)
            for g in sessao.query(GastoFixo).filter(GastoFixo.ano_id == ano.id)
        }

        categoria = (
            sessao.query(Categoria)
            .filter(Categoria.nome == CATEGORIA_PADRAO)
            .one_or_none()
        )

        total = 0
        ignorados: list[str] = []

        for nome_aba in wb.sheetnames:
            ws = wb[nome_aba]
            achado = localizar_tabela(ws)
            if achado is None:
                continue

            primeira, ultima, colunas = achado
            for r in range(primeira, ultima + 1):
                descricao = ws.cell(r, colunas["DESCRICAO"]).value
                valor = para_decimal(ws.cell(r, colunas["VALOR"]).value)
                quando = ws.cell(r, colunas["DATA"]).value

                if not descricao or valor is None:
                    continue  # linha em branco ou a linha de TOTAL
                if valor <= 0:
                    ignorados.append(f"{nome_aba} L{r}: valor {valor} não positivo")
                    continue

                descricao = str(descricao).strip()
                if normalizar(descricao) in existentes:
                    ignorados.append(f"{nome_aba} L{r}: '{descricao}' já cadastrado")
                    continue

                if isinstance(quando, (datetime, date)):
                    dia = quando.day
                else:
                    ignorados.append(
                        f"{nome_aba} L{r}: '{descricao}' sem data, vencimento = dia 1"
                    )
                    dia = 1

                forma = ws.cell(r, colunas["TIPO"]).value if "TIPO" in colunas else None

                sessao.add(
                    GastoFixo(
                        ano_id=ano.id,
                        descricao=descricao,
                        valor=valor,
                        dia_vencimento=dia,
                        forma_pagamento=str(forma).strip() if forma else "",
                        categoria_id=categoria.id if categoria else None,
                    )
                )
                existentes.add(normalizar(descricao))
                total += 1
                print(f"  {descricao:<20} {valor:>8}  vence dia {dia:>2}")

        if simular:
            sessao.rollback()
            print(f"\n[SIMULAÇÃO] {total} gasto(s) fixo(s) seriam importados.")
        else:
            sessao.commit()
            print(f"\n{total} gasto(s) fixo(s) importados para {ano_numero}.")
            print(
                "Todos entraram como pendentes em todos os meses — os lançamentos "
                "de abril já vieram na importação da planilha."
            )

        if ignorados:
            print(f"\n{len(ignorados)} observação(ões):")
            for aviso in ignorados:
                print(f"  - {aviso}")

    finally:
        sessao.close()


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("planilha", type=Path)
    parser.add_argument("--ano", type=int, required=True)
    parser.add_argument("--simular", action="store_true")
    args = parser.parse_args()

    if not args.planilha.is_file():
        print(f"ERRO: arquivo não encontrado: {args.planilha}", file=sys.stderr)
        sys.exit(1)

    criar_tabelas()
    importar(args.planilha, args.ano, args.simular)


if __name__ == "__main__":
    main()
