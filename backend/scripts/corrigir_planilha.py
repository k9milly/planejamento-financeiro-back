"""Gera uma cópia corrigida da planilha original.

Uso:
    python -m scripts.corrigir_planilha "C:/caminho/Planejamento.xlsx"

O arquivo de origem nunca é alterado: a saída vai para um novo arquivo com o
sufixo "- corrigida".

Problemas corrigidos (todos verificados no arquivo original):

1. Linha de agosto do container "Total guardado" somava JUNHO[VALOR] em vez de
   AGOSTO[VALOR] — erro de copiar/colar presente em todas as abas.
2. Título "TOTAL DE MAIO" nas abas de julho a dezembro.
3. A fórmula do total guardado da aba ABRIL não incluía as parcelas de
   "Rendimentos G", ao contrário das outras oito abas: o mesmo dado produzia
   totais diferentes dependendo da aba aberta.
4. O "TOTAL guardado" do mês somava SUMIF(...,"Rendimentos",...), mas esse
   rótulo não existe na planilha (os valores reais são "Rendimentos C" e
   "Rendimentos G"). O SUMIF retornava zero em todas as abas, então rendimento
   nenhum jamais entrou no total guardado.
5. "Rendimentos C" ficava de fora do TOTAL SALDO em abril e de agosto a
   dezembro; só maio, junho e julho o incluíam.
6. O saldo de abertura de cada mês era um número digitado à mão. Agora cada mês
   referencia o fechamento do anterior, então corrigir um lançamento antigo se
   propaga sozinho.
7. Fórmula duplicada e sem rótulo em JUNHO!E9.

ATENÇÃO: openpyxl reescreve o arquivo e pode perder gráficos e imagens. Confira
a cópia antes de descartar o original.
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

import openpyxl

ORDEM = [
    "ABRIL", "MAIO", "JUNHO", "JULHO", "AGOSTO",
    "SETEMBRO", "OUTUBRO", "NOVEMBRO", "DEZEMBRO",
]

# Linhas do container "Total guardado" (A5:B14): uma por mês, sempre nesta ordem.
LINHAS_TOTAL_GUARDADO = {
    "ABRIL": 6, "MAIO": 7, "JUNHO": 8, "JULHO": 9, "AGOSTO": 10,
    "SETEMBRO": 11, "OUTUBRO": 12, "NOVEMBRO": 13, "DEZEMBRO": 14,
}


def formula_total_geral(guardado_inicial: str) -> str:
    """Total guardado acumulado: o que já existia + tudo que entrou - retiradas."""
    guardados = "+".join(f'SUMIF({m}[TIPO],"Guardado",{m}[VALOR])' for m in ORDEM)
    rendimentos = "+".join(
        f'SUMIF({m}[TIPO],"Rendimentos G",{m}[VALOR])' for m in ORDEM
    )
    retiradas = "+".join(f'SUMIF({m}[TIPO],"Retirado",{m}[VALOR])' for m in ORDEM)
    return f"={guardado_inicial}+{guardados}+{rendimentos}-({retiradas})"


def formula_guardado_do_mes(mes: str) -> str:
    return (
        f'=SUMIF({mes}[TIPO],"Guardado",{mes}[VALOR])'
        f'+SUMIF({mes}[TIPO],"Rendimentos G",{mes}[VALOR])'
        f'-SUMIF({mes}[TIPO],"Retirado",{mes}[VALOR])'
    )


def formula_saldo(mes: str, abertura: str) -> str:
    """Entra o recebido, o que rendeu na conta e o que voltou da reserva;
    sai o gasto e o que foi para a reserva."""
    return (
        f"={abertura}"
        f'+SUMIF({mes}[TIPO],"Recebido",{mes}[VALOR])'
        f'+SUMIF({mes}[TIPO],"Rendimentos C",{mes}[VALOR])'
        f'+SUMIF({mes}[TIPO],"Retirado",{mes}[VALOR])'
        f'-SUMIF({mes}[TIPO],"Gasto",{mes}[VALOR])'
        f'-SUMIF({mes}[TIPO],"Guardado",{mes}[VALOR])'
    )


def corrigir(origem: Path, destino: Path, saldo_inicial: str,
             guardado_inicial: str) -> None:
    wb = openpyxl.load_workbook(origem)
    correcoes: list[str] = []

    for indice, nome in enumerate(ORDEM):
        if nome not in wb.sheetnames:
            continue
        ws = wb[nome]

        # (2) Título do container de totais do mês.
        if ws["D3"].value != f"TOTAL DE {nome}":
            correcoes.append(f"{nome}!D3: {ws['D3'].value!r} -> 'TOTAL DE {nome}'")
            ws["D3"] = f"TOTAL DE {nome}"

        # (1) e (3) Total guardado geral: uma única fórmula, igual em toda aba.
        ws["B5"] = formula_total_geral(guardado_inicial)

        # (1) Linha de cada mês dentro do container "Total guardado".
        for mes_linha, linha in LINHAS_TOTAL_GUARDADO.items():
            if mes_linha in wb.sheetnames:
                ws.cell(linha, 2).value = formula_guardado_do_mes(mes_linha)

        # (4) Total guardado do mês, agora somando os rendimentos de fato.
        ws["E7"] = formula_guardado_do_mes(nome)

        # (5) e (6) Saldo: inclui rendimento da conta e encadeia com o mês anterior.
        abertura = saldo_inicial if indice == 0 else f"{ORDEM[indice - 1]}!E8"
        ws["E8"] = formula_saldo(nome, abertura)

        # (7) Sobra de fórmula sem rótulo.
        if nome == "JUNHO" and ws["E9"].value is not None:
            correcoes.append("JUNHO!E9: fórmula duplicada removida")
            ws["E9"] = None

    correcoes.append(
        f"B5/E7/E8 padronizadas em {len(ORDEM)} abas "
        "(rendimentos somados, saldos encadeados)"
    )

    wb.save(destino)

    print(f"Cópia corrigida salva em:\n  {destino}\n")
    print("Correções aplicadas:")
    for item in correcoes:
        print(f"  - {item}")
    print(
        "\nAbra no Excel e confira: openpyxl pode ter descartado gráficos ou "
        "imagens do arquivo original."
    )


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("planilha", type=Path)
    parser.add_argument(
        "--saldo-inicial", default="0.97",
        help="Saldo da conta antes do primeiro mês",
    )
    parser.add_argument(
        "--guardado-inicial", default="7867.36",
        help="Valor guardado antes do primeiro mês",
    )
    parser.add_argument("--saida", type=Path, default=None)
    args = parser.parse_args()

    if not args.planilha.is_file():
        print(f"ERRO: arquivo não encontrado: {args.planilha}", file=sys.stderr)
        sys.exit(1)

    destino = args.saida or args.planilha.with_name(
        f"{args.planilha.stem.strip()} - corrigida.xlsx"
    )
    corrigir(args.planilha, destino, args.saldo_inicial, args.guardado_inicial)


if __name__ == "__main__":
    main()
