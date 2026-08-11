"""Importa a planilha original (.xlsx) para o banco do aplicativo.

Uso:
    python -m scripts.importar_planilha "C:/caminho/Planejamento.xlsx" --ano 2026

A planilha de origem tem uma aba por mês, cada uma com uma tabela de lançamentos
de colunas VALOR / TIPO / CATEGORIA / DATA / OBSERVAÇÕES. As posições das colunas
variam entre abas (uma delas tem uma coluna extra em branco no meio), então o
script localiza o cabeçalho pelo nome em vez de assumir letras fixas.

Mapeamento de tipos (planilha -> aplicativo):
    Recebido        -> entrada
    Gasto           -> saida
    Guardado        -> guardado
    Retirado        -> retirado
    Rendimentos C   -> rendimento (destino: conta)
    Rendimentos G   -> rendimento (destino: guardado)
"""

from __future__ import annotations

import argparse
import sys
import unicodedata
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl

from app.database import SessionLocal, criar_tabelas
from app.models import (
    Ano,
    Categoria,
    DestinoRendimento,
    Lancamento,
    TipoLancamento,
)

MESES = {
    "JANEIRO": 1, "FEVEREIRO": 2, "MARCO": 3, "ABRIL": 4,
    "MAIO": 5, "JUNHO": 6, "JULHO": 7, "AGOSTO": 8,
    "SETEMBRO": 9, "OUTUBRO": 10, "NOVEMBRO": 11, "DEZEMBRO": 12,
}

TIPOS = {
    "RECEBIDO": (TipoLancamento.ENTRADA, None),
    "GASTO": (TipoLancamento.SAIDA, None),
    "GUARDADO": (TipoLancamento.GUARDADO, None),
    "RETIRADO": (TipoLancamento.RETIRADO, None),
    "RENDIMENTOS C": (TipoLancamento.RENDIMENTO, DestinoRendimento.CONTA),
    "RENDIMENTOS G": (TipoLancamento.RENDIMENTO, DestinoRendimento.GUARDADO),
    # Abril usava "Rendimentos" sem distinguir conta de reserva. O único
    # registro assim é de R$ 10,03, e naquele mês a conta tinha R$ 0,97 contra
    # R$ 7.867 guardados — só pode ter rendido na reserva.
    "RENDIMENTOS": (TipoLancamento.RENDIMENTO, DestinoRendimento.GUARDADO),
}

# Cores dos gráficos por categoria; categorias novas recebem cinza.
CORES = {
    "Comida": "#f97316",
    "Lazer": "#8b5cf6",
    "Transporte": "#0ea5e9",
    "Presentes": "#ec4899",
    "Fixo": "#64748b",
    "Outros": "#22c55e",
}


def normalizar(texto: object) -> str:
    """Maiúsculas, sem acento e sem espaços nas pontas.

    A planilha tem 'Estacionamento ' com espaço à direita e acentos
    inconsistentes; comparar sem normalizar criaria categorias duplicadas.
    """
    if texto is None:
        return ""
    sem_acento = unicodedata.normalize("NFKD", str(texto))
    sem_acento = "".join(c for c in sem_acento if not unicodedata.combining(c))
    return sem_acento.strip().upper()


def para_decimal(valor: object) -> Decimal | None:
    if valor is None or isinstance(valor, str) and not valor.strip():
        return None
    try:
        return Decimal(str(valor)).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return None


def para_data(valor: object, ano: int, mes: int) -> date:
    """Converte a célula de data; sem data usável, assume o dia 1º do mês.

    Vários lançamentos da planilha estão sem data — descartá-los perderia
    dinheiro real, então preferimos importar com uma data aproximada.
    """
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    return date(ano, mes, 1)


OBRIGATORIAS = {"VALOR", "TIPO", "DATA"}


def localizar_tabela(ws) -> tuple[int, int, dict[str, int]] | None:
    """Localiza a tabela de lançamentos da aba.

    Devolve (primeira_linha_de_dados, ultima_linha, {coluna: índice}).

    Usa as tabelas nomeadas do Excel (ListObjects) em vez de varrer células.
    É essencial: a aba ABRIL tem seis tabelas lado a lado compartilhando a
    mesma linha de cabeçalho, e três delas têm uma coluna "VALOR" — varrer a
    linha pegaria a coluna errada silenciosamente.
    """
    for tabela in ws.tables.values():
        nomes = [normalizar(c.name) for c in tabela.tableColumns]
        if not OBRIGATORIAS <= set(nomes):
            continue

        col_ini, linha_ini, _, linha_fim = openpyxl.utils.range_boundaries(tabela.ref)
        # `nomes` está na ordem física das colunas, inclusive as sem título
        # (JUNHO tem uma "Column 6" em branco no meio), então o índice bate.
        colunas = {nome: col_ini + i for i, nome in enumerate(nomes)}
        return linha_ini + 1, linha_fim, colunas

    return None


def importar(caminho: Path, ano_numero: int, saldo_conta: Decimal,
             saldo_guardado: Decimal, simular: bool) -> None:
    wb = openpyxl.load_workbook(caminho, data_only=True)
    sessao = SessionLocal()

    try:
        ano = sessao.query(Ano).filter(Ano.ano == ano_numero).one_or_none()
        if ano is None:
            ano = Ano(
                ano=ano_numero,
                saldo_inicial_conta=saldo_conta,
                saldo_inicial_guardado=saldo_guardado,
            )
            sessao.add(ano)
            sessao.flush()
            print(f"Ano {ano_numero} criado.")
        elif sessao.query(Lancamento).filter(Lancamento.ano_id == ano.id).count():
            print(
                f"ERRO: o ano {ano_numero} já tem lançamentos. Apague-os antes de "
                "importar, para não duplicar o histórico.",
                file=sys.stderr,
            )
            sys.exit(1)

        categorias: dict[str, Categoria] = {
            normalizar(c.nome): c for c in sessao.query(Categoria).all()
        }

        total = 0
        ignorados: list[str] = []

        for nome_aba in wb.sheetnames:
            mes = MESES.get(normalizar(nome_aba))
            if mes is None:
                ignorados.append(f"aba '{nome_aba}': não corresponde a um mês")
                continue

            ws = wb[nome_aba]
            achado = localizar_tabela(ws)
            if achado is None:
                ignorados.append(f"aba '{nome_aba}': tabela de lançamentos não achada")
                continue

            primeira, ultima, colunas = achado
            col_valor = colunas["VALOR"]
            col_tipo = colunas["TIPO"]
            col_data = colunas["DATA"]
            col_categoria = colunas.get("CATEGORIA")
            col_obs = colunas.get("OBSERVACOES")

            do_mes = 0
            for r in range(primeira, ultima + 1):
                valor = para_decimal(ws.cell(r, col_valor).value)
                chave_tipo = normalizar(ws.cell(r, col_tipo).value)

                if valor is None or not chave_tipo:
                    continue  # linha vazia da tabela
                if valor <= 0:
                    ignorados.append(f"{nome_aba} linha {r}: valor {valor} não positivo")
                    continue
                if chave_tipo not in TIPOS:
                    ignorados.append(f"{nome_aba} linha {r}: tipo '{chave_tipo}'")
                    continue

                tipo, destino = TIPOS[chave_tipo]

                categoria = None
                if col_categoria and tipo is TipoLancamento.SAIDA:
                    nome_cat = ws.cell(r, col_categoria).value
                    chave = normalizar(nome_cat)
                    if chave:
                        if chave not in categorias:
                            nova = Categoria(
                                nome=str(nome_cat).strip(),
                                cor=CORES.get(str(nome_cat).strip(), "#94a3b8"),
                            )
                            sessao.add(nova)
                            sessao.flush()
                            categorias[chave] = nova
                        categoria = categorias[chave]

                sessao.add(
                    Lancamento(
                        ano_id=ano.id,
                        mes=mes,
                        data=para_data(ws.cell(r, col_data).value, ano_numero, mes),
                        valor=valor,
                        tipo=tipo,
                        destino=destino,
                        categoria_id=categoria.id if categoria else None,
                        descricao=(
                            str(ws.cell(r, col_obs).value).strip()
                            if col_obs and ws.cell(r, col_obs).value
                            else ""
                        ),
                    )
                )
                do_mes += 1

            total += do_mes
            print(f"  {nome_aba:<10} {do_mes:>3} lançamentos")

        if simular:
            sessao.rollback()
            print(f"\n[SIMULAÇÃO] {total} lançamentos seriam importados. Nada foi salvo.")
        else:
            sessao.commit()
            print(f"\n{total} lançamentos importados para {ano_numero}.")

        if ignorados:
            print(f"\n{len(ignorados)} linha(s) ignorada(s):")
            for aviso in ignorados:
                print(f"  - {aviso}")

    finally:
        sessao.close()


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("planilha", type=Path, help="Caminho do arquivo .xlsx")
    parser.add_argument("--ano", type=int, required=True, help="Ano dos dados")
    parser.add_argument(
        "--saldo-conta", type=Decimal, default=Decimal("0"),
        help="Saldo da conta antes do primeiro lançamento",
    )
    parser.add_argument(
        "--saldo-guardado", type=Decimal, default=Decimal("0"),
        help="Valor já guardado antes do primeiro lançamento",
    )
    parser.add_argument(
        "--simular", action="store_true",
        help="Mostra o que seria importado sem gravar nada",
    )
    args = parser.parse_args()

    if not args.planilha.is_file():
        print(f"ERRO: arquivo não encontrado: {args.planilha}", file=sys.stderr)
        sys.exit(1)

    criar_tabelas()
    importar(
        args.planilha, args.ano, args.saldo_conta, args.saldo_guardado, args.simular
    )


if __name__ == "__main__":
    main()
